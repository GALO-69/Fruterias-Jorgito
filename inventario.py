from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import os
import sys

ARCHIVO = "inventario.xlsx"

PRECIOS = {
    "Manzana": 0.50,
    "Banana": 0.30,
    "Naranja": 0.40,
    "Pera": 0.60,
    "Uva": 0.80,
    "Fresa": 0.70,
    "Melon": 1.00,
    "Sandia": 1.20,
    "Piña": 1.50,
    "Mango": 1.80,
    "Kiwi": 1.00,
    "Durazno": 0.90,
    "Cereza": 1.10,
    "Frambuesa": 1.30,
    "Arandano": 1.40,
    "Ciruela": 0.75,
}

CANTIDAD_MAXIMA = 100
CANTIDAD_MINIMA = 25


# ----------------------------
# UTILIDADES
# ----------------------------

def crear_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.append(["Fruta", "Cantidad", "Precio"])
    guardar_excel(wb)


def abrir_excel():
    try:
        if not os.path.exists(ARCHIVO):
            crear_excel()
        return load_workbook(ARCHIVO)
    except InvalidFileException:
        print("El archivo Excel esta corrupto.")
        sys.exit(1)
    except Exception as e:
        print(f"Error inesperado al abrir el archivo: {e}")
        sys.exit(1)


def guardar_excel(wb):
    try:
        wb.save(ARCHIVO)
    except PermissionError:
        print("No se puede guardar el archivo. Cierre Excel e intente nuevamente.")
    except Exception as e:
        print(f"Error inesperado al guardar el archivo: {e}")


def buscar_fruta(ws, fruta):
    for fila in ws.iter_rows(min_row=2, values_only=False):
        if fila[0].value == fruta:
            return fila
    return None


def solicitar_entero_positivo(mensaje):
    try:
        valor = int(input(mensaje))
        if valor <= 0:
            raise ValueError
        return valor
    except ValueError:
        print("Debe ingresar un numero entero positivo.")
        return None


# ----------------------------
# LOGICA DE NEGOCIO
# ----------------------------

def agregar_fruta():
    wb = abrir_excel()
    ws = wb["Inventario"]

    fruta = input("Nombre de la fruta: ").strip().capitalize()

    if fruta not in PRECIOS:
        print("Fruta no permitida.")
        return

    cantidad = solicitar_entero_positivo("Cantidad a agregar: ")
    if cantidad is None:
        return

    fila = buscar_fruta(ws, fruta)

    if fila:
        cantidad_actual = fila[1].value
        nueva_cantidad = cantidad_actual + cantidad

        if nueva_cantidad > CANTIDAD_MAXIMA:
            print("La cantidad supera el maximo permitido.")
            return

        fila[1].value = nueva_cantidad
    else:
        if cantidad < CANTIDAD_MINIMA:
            print("La cantidad inicial debe ser al menos 25.")
            return

        if cantidad > CANTIDAD_MAXIMA:
            print("La cantidad supera el maximo permitido.")
            return

        ws.append([fruta, cantidad, PRECIOS[fruta]])

    guardar_excel(wb)
    print("Inventario actualizado correctamente.")


def vender_fruta():
    wb = abrir_excel()
    ws = wb["Inventario"]

    fruta = input("Fruta vendida: ").strip().capitalize()
    fila = buscar_fruta(ws, fruta)

    if not fila:
        print("La fruta no existe en el inventario.")
        return

    cantidad = solicitar_entero_positivo("Cantidad vendida: ")
    if cantidad is None:
        return

    cantidad_actual = fila[1].value
    nueva_cantidad = cantidad_actual - cantidad

    if nueva_cantidad < CANTIDAD_MINIMA:
        print("No se puede reducir por debajo del minimo permitido.")
        return

    fila[1].value = nueva_cantidad
    guardar_excel(wb)
    print("Venta registrada correctamente.")


def mostrar_inventario():
    wb = abrir_excel()
    ws = wb["Inventario"]

    datos = list(ws.iter_rows(values_only=True))

    if len(datos) <= 1:
        print("El inventario esta vacio.")
        return

    encabezados = datos[0]

    print("\nINVENTARIO ACTUAL\n")
    print(f"{encabezados[0]:<12} | {encabezados[1]:<10} | {encabezados[2]:<10}")
    print("-" * 40)

    for fila in datos[1:]:
        print(f"{fila[0]:<12} | {fila[1]:<10} | {fila[2]:<10}")

    print("-" * 40)


# ----------------------------
# MENU PRINCIPAL
# ----------------------------

def menu():
    while True:
        print("\nSISTEMA DE INVENTARIO")
        print("1. Agregar fruta")
        print("2. Vender fruta")
        print("3. Mostrar inventario")
        print("4. Salir")

        opcion = input("Seleccione una opcion: ").strip()

        match opcion:
            case "1":
                agregar_fruta()
            case "2":
                vender_fruta()
            case "3":
                mostrar_inventario()
            case "4":
                print("Programa finalizado.")
                break
            case _:
                print("Opcion invalida.")


if __name__ == "__main__":
    menu()